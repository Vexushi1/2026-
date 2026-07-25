from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from scipy.stats import spearmanr


随机种子 = 2026
np.random.seed(随机种子)

项目根目录 = Path(__file__).resolve().parent
输入文件 = 项目根目录 / "A题数据集.csv"
结果目录 = 项目根目录 / "结果数据表"
输出文件 = 结果目录 / "数据审计与预处理结果.xlsx"

目标字段 = ["Early_Waker", "Health_Score", "Wellness_Category"]
时间字段 = ["Wake_Up_Time", "Sleep_Time"]
标识字段 = ["Person_ID"]

预期字段 = [
    "Person_ID", "Age", "Gender", "Height_cm", "Weight_kg", "BMI", "Country",
    "Occupation", "Marital_Status", "Wake_Up_Time", "Sleep_Time",
    "Sleep_Duration_Hours", "Sleep_Quality_Score", "Number_of_Night_Awakenings",
    "Weekend_Sleep_Difference_Hours", "Nap_Frequency_Per_Week",
    "Screen_Time_Before_Bed_Hours", "Exercise_Frequency_Per_Week",
    "Exercise_Duration_Minutes", "Exercise_Type", "Daily_Steps", "Morning_Workout",
    "Workout_Intensity", "Gym_Member", "Daily_Calorie_Intake", "Water_Intake_Liters",
    "Fruit_Intake_Per_Day", "Vegetable_Intake_Per_Day", "Protein_Intake_Grams",
    "Sugary_Drinks_Per_Week", "Fast_Food_Meals_Per_Week",
    "Breakfast_Regularity_Score", "Smoking_Status", "Alcohol_Consumption",
    "Stress_Level", "Working_Hours_Per_Day", "Sitting_Hours_Per_Day",
    "Outdoor_Time_Hours", "Social_Interaction_Score", "Meditation_Practice",
    "Resting_Heart_Rate", "Systolic_BP", "Diastolic_BP", "Cholesterol_Level",
    "Blood_Sugar_Level", "Energy_Level_Score", "Fatigue_Level_Score",
    "Immune_Health_Score", "Mood_Score", "Anxiety_Score", "Depression_Risk_Score",
    "Productivity_Score", "Focus_Concentration_Score", "Life_Satisfaction_Score",
    "Obesity_Risk", "Hypertension_Risk", "Diabetes_Risk", "Cardiovascular_Risk",
    "Sleep_Disorder_Risk", "Health_Score", "Fitness_Level", "Healthy_Aging_Score",
    "Wellness_Category", "Early_Waker",
]

硬范围 = {
    "Age": (0, 120),
    "Height_cm": (100, 250),
    "Weight_kg": (20, 400),
    "BMI": (5, 100),
    "Sleep_Duration_Hours": (0, 24),
    "Number_of_Night_Awakenings": (0, 30),
    "Weekend_Sleep_Difference_Hours": (-12, 12),
    "Nap_Frequency_Per_Week": (0, 21),
    "Screen_Time_Before_Bed_Hours": (0, 24),
    "Exercise_Frequency_Per_Week": (0, 14),
    "Exercise_Duration_Minutes": (0, 600),
    "Daily_Steps": (0, 100000),
    "Daily_Calorie_Intake": (0, 10000),
    "Water_Intake_Liters": (0, 15),
    "Fruit_Intake_Per_Day": (0, 20),
    "Vegetable_Intake_Per_Day": (0, 20),
    "Protein_Intake_Grams": (0, 500),
    "Sugary_Drinks_Per_Week": (0, 50),
    "Fast_Food_Meals_Per_Week": (0, 50),
    "Working_Hours_Per_Day": (0, 24),
    "Sitting_Hours_Per_Day": (0, 24),
    "Outdoor_Time_Hours": (0, 24),
    "Resting_Heart_Rate": (20, 250),
    "Systolic_BP": (50, 300),
    "Diastolic_BP": (20, 200),
    "Cholesterol_Level": (0, 500),
    "Blood_Sugar_Level": (0, 1000),
}


@dataclass(frozen=True)
class 审计结果:
    数据: pd.DataFrame
    工作表: dict[str, pd.DataFrame]


def 读取原始数据(path: Path = 输入文件) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"未找到数据文件：{path}")

    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            df = pd.read_csv(path, encoding=encoding, keep_default_na=False)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise RuntimeError("CSV编码无法识别") from last_error

    df.columns = [str(col).strip() for col in df.columns]
    object_cols = df.select_dtypes(include="object").columns
    for col in object_cols:
        df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)
        df.loc[df[col].eq(""), col] = np.nan

    缺失字段 = [col for col in 预期字段 if col not in df.columns]
    多余字段 = [col for col in df.columns if col not in 预期字段]
    if 缺失字段:
        raise ValueError(f"缺少预期字段：{缺失字段}")
    if 多余字段:
        print(f"警告：发现题目字段表之外的列：{多余字段}")
    return df


def 时间转分钟(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, format="%H:%M", errors="coerce")
    return parsed.dt.hour * 60 + parsed.dt.minute


def 字段字典(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for col in df.columns:
        if col in 标识字段:
            category = "标识字段"
            treatment = "仅作结果关联，不进入模型"
        elif col in 时间字段:
            category = "时间字段"
            treatment = "转换为分钟；模型分支按需增加周期编码"
        elif col in 目标字段:
            category = "目标字段"
            treatment = "按任务隔离，不作为同任务普通输入"
        elif pd.api.types.is_numeric_dtype(df[col]):
            category = "数值字段"
            treatment = "合法范围和IQR检查；按模型需要缩放"
        else:
            category = "类别字段"
            treatment = "保留显式None；训练折内编码或CatBoost原生输入"

        rows.append({
            "字段": col,
            "数据类型": str(df[col].dtype),
            "变量类别": category,
            "非空数": int(df[col].notna().sum()),
            "唯一值数": int(df[col].nunique(dropna=True)),
            "当前处理原则": treatment,
        })
    return pd.DataFrame(rows)


def 完整性检查(df: pd.DataFrame) -> pd.DataFrame:
    numeric_cols = df.select_dtypes(include=np.number).columns
    object_cols = df.select_dtypes(exclude=np.number).columns
    illegal_numeric = 0
    for col in numeric_cols:
        illegal_numeric += int(pd.to_numeric(df[col], errors="coerce").isna().sum() - df[col].isna().sum())

    checks = [
        ("样本数", len(df), "应为10000；变化时重新同步框架"),
        ("字段数", df.shape[1], f"预期{len(预期字段)}列"),
        ("完全重复行", int(df.duplicated().sum()), "应为0"),
        ("重复Person_ID", int(df["Person_ID"].duplicated().sum()), "应为0"),
        ("真实空值单元格", int(df.isna().sum().sum()), "不含字符串None"),
        ("目标字段空值", int(df[目标字段].isna().sum().sum()), "应为0"),
        ("数值列非法字符串", illegal_numeric, "应为0"),
        ("类别字段数量", len(object_cols), "用于编码策略"),
        ("数值字段数量", len(numeric_cols), "用于范围与描述统计"),
    ]
    return pd.DataFrame(checks, columns=["检验项目", "结果", "判定说明"])


def 特殊类别检查(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for col in df.select_dtypes(include="object").columns:
        values = df[col].astype("string")
        none_mask = values.str.casefold().eq("none")
        if none_mask.any():
            extra = ""
            if col in {"Exercise_Type", "Workout_Intensity"}:
                zero_freq = pd.to_numeric(df["Exercise_Frequency_Per_Week"], errors="coerce").eq(0)
                extra = f"与锻炼频率为0重合率={float((none_mask & zero_freq).sum() / none_mask.sum()):.4f}"
            rows.append({
                "字段": col,
                "None数量": int(none_mask.sum()),
                "None比例": float(none_mask.mean()),
                "语义处理": "作为显式类别保留，不替换为缺失值",
                "联合检查": extra,
            })
    if not rows:
        rows.append({"字段": "无", "None数量": 0, "None比例": 0.0,
                     "语义处理": "未发现字符串None", "联合检查": ""})
    return pd.DataFrame(rows)


def 描述统计(df: pd.DataFrame) -> pd.DataFrame:
    numeric = df.select_dtypes(include=np.number)
    desc = numeric.describe(percentiles=[0.01, 0.25, 0.5, 0.75, 0.99]).T
    desc["缺失数"] = numeric.isna().sum()
    desc["偏度"] = numeric.skew(numeric_only=True)
    desc["峰度"] = numeric.kurt(numeric_only=True)
    desc = desc.reset_index(names="字段")
    rename = {
        "count": "非空数", "mean": "均值", "std": "标准差", "min": "最小值",
        "1%": "P01", "25%": "Q1", "50%": "中位数", "75%": "Q3",
        "99%": "P99", "max": "最大值",
    }
    return desc.rename(columns=rename)


def 异常与范围检查(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    numeric = df.select_dtypes(include=np.number)
    for col in numeric.columns:
        values = pd.to_numeric(numeric[col], errors="coerce").dropna()
        if values.empty:
            continue
        q1, q3 = values.quantile([0.25, 0.75])
        iqr = q3 - q1
        low_iqr, high_iqr = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        iqr_count = int(((values < low_iqr) | (values > high_iqr)).sum())
        hard_low, hard_high = 硬范围.get(col, (np.nan, np.nan))
        hard_count = 0
        if not np.isnan(hard_low):
            hard_count = int(((values < hard_low) | (values > hard_high)).sum())
        rows.append({
            "字段": col,
            "最小值": float(values.min()),
            "最大值": float(values.max()),
            "IQR下界": float(low_iqr),
            "IQR上界": float(high_iqr),
            "IQR标记数": iqr_count,
            "硬下界": hard_low,
            "硬上界": hard_high,
            "硬边界违反数": hard_count,
            "处理结论": "硬边界违反需核验；仅IQR离群默认保留",
        })

    bp_violation = int((pd.to_numeric(df["Systolic_BP"], errors="coerce") <
                        pd.to_numeric(df["Diastolic_BP"], errors="coerce")).sum())
    rows.append({
        "字段": "Systolic_BP<Diastolic_BP",
        "最小值": np.nan, "最大值": np.nan, "IQR下界": np.nan, "IQR上界": np.nan,
        "IQR标记数": np.nan, "硬下界": np.nan, "硬上界": np.nan,
        "硬边界违反数": bp_violation,
        "处理结论": "若大于0需逐行核验血压字段",
    })
    return pd.DataFrame(rows)


def 最优二分类阈值(x: pd.Series, y: pd.Series) -> tuple[float, int, float]:
    valid = x.notna() & y.notna()
    x_arr = x[valid].astype(float).to_numpy()
    y_arr = y[valid].astype(int).to_numpy()
    order = np.argsort(x_arr)
    x_sorted, y_sorted = x_arr[order], y_arr[order]

    unique_x, first_idx = np.unique(x_sorted, return_index=True)
    candidates = np.r_[unique_x[0] - 0.5,
                       (unique_x[:-1] + unique_x[1:]) / 2,
                       unique_x[-1] + 0.5]
    total_pos = y_sorted.sum()
    cum_pos = np.r_[0, np.cumsum(y_sorted)]
    cum_neg = np.r_[0, np.cumsum(1 - y_sorted)]

    split_indices = np.searchsorted(x_sorted, candidates, side="left")
    # 规则：x<threshold预测1，左侧负类+右侧正类为错误
    errors = cum_neg[split_indices] + (total_pos - cum_pos[split_indices])
    best = int(np.argmin(errors))
    threshold = float(candidates[best])
    error_count = int(errors[best])
    accuracy = 1.0 - error_count / len(x_sorted)
    return threshold, error_count, float(accuracy)


def 最优有序切点(score: pd.Series, labels: pd.Series) -> tuple[list[str], list[float], int, float]:
    valid = score.notna() & labels.notna()
    x = score[valid].astype(float)
    y = labels[valid].astype(str)
    ordered_classes = y.groupby(y).apply(lambda s: x.loc[s.index].median()).sort_values().index.tolist()
    class_to_idx = {name: idx for idx, name in enumerate(ordered_classes)}

    table = pd.crosstab(x, y).reindex(columns=ordered_classes, fill_value=0).sort_index()
    unique_scores = table.index.to_numpy(dtype=float)
    counts = table.to_numpy(dtype=int)
    m, k = counts.shape
    prefix = np.vstack([np.zeros((1, k), dtype=int), np.cumsum(counts, axis=0)])

    dp = np.full((k + 1, m + 1), np.inf)
    prev = np.full((k + 1, m + 1), -1, dtype=int)
    dp[0, 0] = 0.0

    def segment_cost(left: int, right: int, cls: int) -> int:
        segment = prefix[right] - prefix[left]
        return int(segment.sum() - segment[cls])

    for cls in range(1, k + 1):
        for right in range(cls, m + 1):
            best_cost = np.inf
            best_left = -1
            for left in range(cls - 1, right):
                cost = dp[cls - 1, left] + segment_cost(left, right, cls - 1)
                if cost < best_cost:
                    best_cost, best_left = cost, left
            dp[cls, right], prev[cls, right] = best_cost, best_left

    cuts: list[int] = []
    right = m
    for cls in range(k, 0, -1):
        left = prev[cls, right]
        cuts.append(left)
        right = left
    cuts = sorted(cuts[1:])

    thresholds: list[float] = []
    for cut in cuts:
        thresholds.append(float((unique_scores[cut - 1] + unique_scores[cut]) / 2))
    errors = int(dp[k, m])
    accuracy = 1.0 - errors / int(valid.sum())
    return ordered_classes, thresholds, errors, float(accuracy)


def 派生关系检查(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    height_m = pd.to_numeric(df["Height_cm"], errors="coerce") / 100
    bmi_calc = pd.to_numeric(df["Weight_kg"], errors="coerce") / height_m.pow(2)
    bmi_error = (pd.to_numeric(df["BMI"], errors="coerce") - bmi_calc).abs()
    rows.append({
        "关系": "BMI=体重/身高²",
        "样本数": int(bmi_error.notna().sum()),
        "一致率/相关": float((bmi_error <= 0.15).mean()),
        "最大误差": float(bmi_error.max()),
        "违反数": int((bmi_error > 0.15).sum()),
        "当前结论": "若误差仅为舍入，BMI作为派生变量保留并标注冗余",
    })

    wake = 时间转分钟(df["Wake_Up_Time"])
    sleep = 时间转分钟(df["Sleep_Time"])
    duration_calc = ((wake - sleep) % 1440) / 60
    duration_error = (pd.to_numeric(df["Sleep_Duration_Hours"], errors="coerce") - duration_calc).abs()
    rows.append({
        "关系": "睡眠时长=(起床-入睡) mod 24h",
        "样本数": int(duration_error.notna().sum()),
        "一致率/相关": float((duration_error <= 0.11).mean()),
        "最大误差": float(duration_error.max()),
        "违反数": int((duration_error > 0.11).sum()),
        "当前结论": "用于确认时间转换和字段派生关系",
    })

    early = df["Early_Waker"].astype(str).str.casefold().isin({"yes", "1", "true"}).astype(int)
    threshold, errors, accuracy = 最优二分类阈值(wake, early)
    rows.append({
        "关系": "Early_Waker与Wake_Up_Time阈值",
        "样本数": int(wake.notna().sum()),
        "一致率/相关": accuracy,
        "最大误差": threshold,
        "违反数": errors,
        "当前结论": "阈值以分钟表示；正式模型在训练折内估计",
    })

    fitness = df["Fitness_Level"].astype("string")
    wellness = df["Wellness_Category"].astype("string")
    same = fitness.eq(wellness) & fitness.notna() & wellness.notna()
    rows.append({
        "关系": "Fitness_Level与Wellness_Category",
        "样本数": int((fitness.notna() & wellness.notna()).sum()),
        "一致率/相关": float(same.mean()),
        "最大误差": np.nan,
        "违反数": int((~same & fitness.notna() & wellness.notna()).sum()),
        "当前结论": "一致关系作为问题三结构先验，仍需设置冲突回退",
    })

    ordered_classes, cuts, cut_errors, cut_accuracy = 最优有序切点(
        pd.to_numeric(df["Health_Score"], errors="coerce"), df["Wellness_Category"]
    )
    rows.append({
        "关系": "Health_Score对Wellness_Category的最优有序分段",
        "样本数": int(df["Health_Score"].notna().sum()),
        "一致率/相关": cut_accuracy,
        "最大误差": ", ".join(f"{x:.3f}" for x in cuts),
        "违反数": cut_errors,
        "当前结论": "类别顺序=" + "<".join(ordered_classes),
    })

    risk_specs = [
        ("BMI", "Obesity_Risk"),
        ("Systolic_BP", "Hypertension_Risk"),
        ("Blood_Sugar_Level", "Diabetes_Risk"),
        ("Sleep_Quality_Score", "Sleep_Disorder_Risk"),
    ]
    for numeric_col, risk_col in risk_specs:
        tmp = df[[numeric_col, risk_col]].dropna().copy()
        medians = tmp.groupby(risk_col)[numeric_col].median().sort_values()
        order_map = {name: idx for idx, name in enumerate(medians.index)}
        encoded = tmp[risk_col].map(order_map)
        corr, _ = spearmanr(pd.to_numeric(tmp[numeric_col], errors="coerce"), encoded)
        rows.append({
            "关系": f"{numeric_col}与{risk_col}",
            "样本数": len(tmp),
            "一致率/相关": float(corr),
            "最大误差": np.nan,
            "违反数": np.nan,
            "当前结论": "类别中位数顺序：" + "<".join(map(str, medians.index)),
        })

    return pd.DataFrame(rows)


def 目标结构(df: pd.DataFrame) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for target in ("Early_Waker", "Wellness_Category", "Fitness_Level"):
        counts = df[target].value_counts(dropna=False).rename_axis("类别").reset_index(name="样本数")
        counts["比例"] = counts["样本数"] / len(df)
        counts.insert(0, "目标字段", target)
        frames.append(counts)

    score = pd.to_numeric(df["Health_Score"], errors="coerce")
    q = score.quantile([0.25, 0.5, 0.75]).to_numpy()
    quantile_labels = pd.cut(score, [-np.inf, q[0], q[1], q[2], np.inf],
                             labels=["Poor", "Average", "Good", "Excellent"],
                             include_lowest=True, duplicates="drop")
    quantile_counts = quantile_labels.value_counts(sort=False).rename_axis("类别").reset_index(name="样本数")
    quantile_counts["比例"] = quantile_counts["样本数"] / quantile_counts["样本数"].sum()
    quantile_counts.insert(0, "目标字段", "Health_Score四分位候选")
    frames.append(quantile_counts)

    return pd.concat(frames, ignore_index=True)


def 衍生特征说明() -> pd.DataFrame:
    rows = [
        ("Wake_Min", "Wake_Up_Time", "$60h+m$", "问题一阈值；问题二、三周期编码基础"),
        ("Wake_Sin/Wake_Cos", "Wake_Min", "$\sin(2\pi t/1440),\cos(2\pi t/1440)$", "保持时间周期邻近性"),
        ("Sleep_Sin/Sleep_Cos", "Sleep_Time", "$\sin(2\pi t/1440),\cos(2\pi t/1440)$", "保持入睡时间周期邻近性"),
        ("Exercise_Load", "锻炼频率、单次时长", "$f\times d$", "每周总运动负荷"),
        ("Sit_Work_Ratio", "久坐时长、工作时长", "$s/(w+\varepsilon)$", "工作结构中的久坐暴露"),
        ("Plant_Intake", "水果、蔬菜", "$fruit+vegetable$", "植物性食物摄入"),
        ("Unhealthy_Diet_Load", "含糖饮料、快餐", "$sugary+fastfood$", "不健康饮食负荷"),
        ("Mental_Load", "压力、焦虑、抑郁风险、情绪", "训练折标准化后线性组合", "避免原始量纲直接相加"),
        ("BP_Pulse_Pressure", "收缩压、舒张压", "$SBP-DBP$", "脉压差"),
    ]
    return pd.DataFrame(rows, columns=["衍生特征", "来源字段", "计算方式", "现实含义"])


def 特征准入矩阵(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for col in df.columns:
        q1 = q2 = q3 = "普通输入候选"
        reason = "经审计后按模型分支编码"
        if col == "Person_ID":
            q1 = q2 = q3 = "排除"
            reason = "仅为关联键"
        elif col == "Early_Waker":
            q1, q2, q3 = "目标", "排除", "排除"
            reason = "问题一目标"
        elif col == "Health_Score":
            q1, q2, q3 = "普通输入候选", "目标基础", "结构先验"
            reason = "问题二连续目标；问题三健康评分先验"
        elif col == "Wellness_Category":
            q1, q2, q3 = "普通输入候选", "结构变量待审计", "目标"
            reason = "问题三目标；问题二是否准入由派生关系和折内验证决定"
        elif col == "Fitness_Level":
            q1, q2, q3 = "普通输入候选", "结构变量待审计", "结构先验"
            reason = "与综合健康类别的一致性需量化"
        elif col == "Wake_Up_Time":
            q1, q2, q3 = "结构先验", "周期输入", "周期输入"
            reason = "问题一时间边界；其他问题按周期编码"
        elif col == "Sleep_Time":
            q1, q2, q3 = "辅助输入", "周期输入", "周期输入"
            reason = "时间字段，不直接按字符串建模"
        elif col in {"BMI", "Obesity_Risk", "Hypertension_Risk", "Diabetes_Risk",
                     "Cardiovascular_Risk", "Sleep_Disorder_Risk", "Healthy_Aging_Score"}:
            q1 = q2 = q3 = "综合/派生输入"
            reason = "保留但记录生成关系；消融检验其必要性"

        rows.append({"字段": col, "问题一": q1, "问题二": q2, "问题三": q3, "准入依据": reason})
    return pd.DataFrame(rows)


def 管道检查说明() -> pd.DataFrame:
    rows = [
        (1, "先划分外层训练折和验证折", "标签分层，仅保留验证折用于折外预测", "必须"),
        (2, "训练折内拟合预处理器", "类别字典、标准化参数、异常边界均不得读取验证折", "必须"),
        (3, "训练折内估计问题一阈值", "验证折只用于评价", "必须"),
        (4, "训练折内估计问题二等级切点", "同一组切点标记训练折与验证折", "必须"),
        (5, "训练折内选择超参数与融合权重", "需要时使用内层交叉验证", "必须"),
        (6, "输出OOF概率和标签", "每个样本仅由未见过该样本的模型预测", "必须"),
        (7, "全量训练最终模型", "仅在模型口径冻结后执行", "后续"),
    ]
    return pd.DataFrame(rows, columns=["步骤", "操作", "防泄漏要求", "状态"])


def 执行审计(df: pd.DataFrame) -> 审计结果:
    data_overview = pd.DataFrame([
        ("样本数", len(df), "行"),
        ("字段数", df.shape[1], "列"),
        ("数值字段数", len(df.select_dtypes(include=np.number).columns), "列"),
        ("类别/时间字段数", len(df.select_dtypes(exclude=np.number).columns), "列"),
        ("目标字段", ", ".join(目标字段), ""),
        ("关联键", "Person_ID", ""),
        ("随机种子", 随机种子, ""),
    ], columns=["指标", "结果", "单位或说明"])

    sheets = {
        "数据概览": data_overview,
        "字段字典": 字段字典(df),
        "完整性检查": 完整性检查(df),
        "特殊类别": 特殊类别检查(df),
        "描述统计": 描述统计(df),
        "异常与范围": 异常与范围检查(df),
        "派生关系": 派生关系检查(df),
        "目标结构": 目标结构(df),
        "衍生特征": 衍生特征说明(),
        "特征准入": 特征准入矩阵(df),
        "管道检查": 管道检查说明(),
    }
    return 审计结果(数据=df, 工作表=sheets)


def 设置工作簿样式(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="173B5E")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    body_font = Font(name="Microsoft YaHei", size=10)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 38)
    wb.save(path)


def 保存结果(result: 审计结果, path: Path = 输出文件) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, table in result.工作表.items():
            if table.empty:
                table = pd.DataFrame({"适用性说明": ["本工作表当前无记录"]})
            table.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    设置工作簿样式(path)


def main() -> None:
    df = 读取原始数据()
    result = 执行审计(df)
    保存结果(result)

    summary = result.工作表["完整性检查"].set_index("检验项目")["结果"]
    print(f"数据审计完成：{输出文件}")
    print(f"样本数={summary['样本数']}，字段数={summary['字段数']}，重复行={summary['完全重复行']}")
    print("下一步：根据审计结果冻结异常处理、问题二等级切点和三问特征准入矩阵。")


if __name__ == "__main__":
    main()
